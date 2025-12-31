import math
import os
from collections import defaultdict
from datetime import datetime, timedelta
import traceback
from urllib.parse import quote, urljoin
import shutil
import json
from threading import Thread
import geopandas as gpd
import numpy as np
import pandas as pd
import yagmail
from dotenv import load_dotenv
from flask import Flask, jsonify, make_response, request, g
from psycopg2.extras import execute_values
from flask_jwt_extended import (
    JWTManager,
    create_access_token,
    create_refresh_token,
    decode_token,
    get_jwt,
    get_jwt_identity,
    jwt_required,
    verify_jwt_in_request
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from psycopg2 import extras
from psycopg2.extras import RealDictCursor
from psycopg2.extras import DictCursor
from shapely import from_wkt, wkt
from shapely.affinity import rotate, translate
from shapely.geometry import MultiPolygon, Polygon, box
from shapely.ops import split, unary_union
from help_func import format_hazard_records, format_device_name, get_device_label
from db import get_db_conn, release_db_conn
import requests
from apscheduler.schedulers.background import BackgroundScheduler
load_dotenv()

whatsapp_auth_key = (os.environ.get("whatsapp_authkey"),)
EMAIL = os.environ.get("EMAIL")
PASSWORD = os.environ.get("PASSOWRD")
RECIVERS = os.environ.get("CC_RECEIVERS")

from flask_cors import CORS, cross_origin

app = Flask(__name__)
CORS(app)

# JWT Configuration
app.config["JWT_SECRET_KEY"] = "t7knf74gjsjv6ckj3$go#Glw64"
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=1)
app.config["JWT_TOKEN_LOCATION"] = ["headers"]

jwt = JWTManager(app)

# Scheduler
scheduler = BackgroundScheduler(daemon=True)
def run_every_minute():
    # SAVE NEW SESSION
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE weatherdata.licensed_user_auth
                SET online_status = 'offline'
                WHERE userid in (SELECT user_id FROM weatherdata.user_sessions WHERE expires_at < NOW() OR last_request < NOW() - INTERVAL '10 minutes')
                """
            )

            cur.execute(
                """
                UPDATE weatherdata.weather_user_activity_log
                SET logout_time = NOW()
                WHERE id in (SELECT log_id FROM weatherdata.user_sessions WHERE expires_at < NOW() OR last_request < NOW() - INTERVAL '10 minutes')
                """
            )

            cur.execute(
                """
                DELETE FROM weatherdata.user_sessions WHERE expires_at < NOW() OR last_request < NOW() - INTERVAL '10 minutes'
                """
            )
            conn.commit()
    finally:
        release_db_conn(conn) 

scheduler.add_job(
    run_every_minute,
    trigger="interval",
    minutes=1,
    id="minute_job",
    replace_existing=True
)
scheduler.start()

@jwt.token_in_blocklist_loader
def check_if_token_revoked(jwt_header, jwt_payload):
    jti = jwt_payload.get("jti")
    user_id = jwt_payload.get("sub")

    if not user_id or not jti:
        return True  # block invalid token

    conn = get_db_conn()
    try:
        with conn.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(
                """
                SELECT jti
                FROM weatherdata.user_sessions
                WHERE user_id = %s
                """,
                (user_id,)
            )
            session = cur.fetchone()

            # No active session → token revoked
            if not session:
                return True

            # If JTI does not match → logged in elsewhere
            return session["jti"] != jti

    finally:
        release_db_conn(conn)

# --- JWT error handlers ----
@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    return jsonify({"msg": "Token has expired"}), 401

@jwt.invalid_token_loader
def invalid_token_callback(error_string):
    return jsonify({"msg": f"Invalid token: {error_string}"}), 401

@jwt.unauthorized_loader
def missing_token_callback(error_string):
    return jsonify({"msg": f"Missing token: {error_string}"}), 401

@jwt.revoked_token_loader
def revoked_token_callback(jwt_header, jwt_payload):
    return jsonify({"msg": "Token has been revoked"}), 401

@app.before_request
def before():
    if request.method == "OPTIONS":
        return
    try:
        verify_jwt_in_request(optional=True)
        g.user_id = get_jwt_identity()
    except:
        g.user_id = None
    if not g.user_id:
        return
    async_update(g.user_id)
            
def update_last_request_db(user_id):
    conn = None
    try:
        conn = get_db_conn()
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE weatherdata.user_sessions
                SET last_request = NOW()
                WHERE user_id = %s
                """,
                (user_id,)
            )
        conn.commit()
    except Exception as e:
        print("Async DB error:", e)
    finally:
        if conn:
            release_db_conn(conn)
            
def async_update(user_id):
    Thread(
        target=update_last_request_db,
        args=(user_id,),
        daemon=True  # auto-kill with app
    ).start()
    
# mark DB online/offline
def mark_user_online_offline(userid: str, online: bool):
    conn = None
    try:
        conn = get_db_conn()
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE weatherdata.licensed_user_auth SET online_status = %s WHERE userid = %s",
                ("online" if online else "offline", userid),
            )
            conn.commit()
    except Exception:
        if conn:
            conn.rollback()
    finally:
        release_db_conn(conn)

def insert_activity_log(userid: str, username: str, device: str, name:str):
    """Insert a login entry into activity log and return inserted id if available."""
    conn = get_db_conn()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            login_time = datetime.now().strftime("%d %b %Y %H:%M")
            cur.execute(
                """
                INSERT INTO weatherdata.weather_user_activity_log (userid, username, loggedin_device, login_time, name)
                VALUES (%s, %s, %s, NOW(), %s)
                RETURNING id
                """,
                (userid, username, device, name),
            )
            res = cur.fetchone()
            conn.commit()
            if res:
                return res.get("id")
    except Exception:
        if conn:
            conn.rollback()
    finally:
        release_db_conn(conn)
    return None

def update_activity_logout(log_id):
    conn = None
    try:
        conn = get_db_conn()
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE weatherdata.weather_user_activity_log SET logout_time = NOW() WHERE id = %s",
                (log_id,),
            )
            conn.commit()
    except Exception:
        if conn:
            conn.rollback()
    finally:
        release_db_conn(conn)

# Routes
@app.route("/")
@cross_origin("*")
def index():
    return jsonify({"msg": "Weather API running....."}), 200

@app.route("/userLogin", methods=["POST"])
@cross_origin()
def user_login():
    conn = None
    try:
        data = request.get_json() or {}
        username = data.get("username")
        userpassword = data.get("userpassword")
        force_login = data.get("force_login", False)

        if not username or not userpassword:
            return jsonify({"msg": "username and userpassword required"}), 400

        conn = get_db_conn()

        # Fetch user
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT *
                FROM weatherdata.licensed_user_auth
                WHERE (username = %s OR mail = %s)
                """,
                (username, username),
            )
            user = cur.fetchone()

        if not user:
            return jsonify({"msg": "User not found"}), 404

        if user.get("status") != "active":
            return jsonify({"msg": "User account not active"}), 403

        if user.get("password") != userpassword:
            return jsonify({"msg": "Invalid Password"}), 401

        userid = user["userid"]
        # device = request.headers.get("User-Agent", "Unknown")
        device = get_device_label(request.headers.get("User-Agent", "Unknown"))

        # CHECK EXISTING SESSION
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT jti, log_id
                FROM weatherdata.user_sessions
                WHERE user_id = %s AND expires_at > NOW()
                """,
                (userid,),
            )
            existing_session = cur.fetchone()

        # User already logged in 
        if existing_session and not force_login:
            log_id = existing_session.get('log_id')
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                        select name, userid, login_time, loggedin_device from weatherdata.weather_user_activity_log
                        where id = %s
                    """,
                    (log_id,),
                )
                recent_log = cur.fetchone()
            
            return jsonify({
                "data": {"name":recent_log.get('name'), "userid":recent_log.get('userid'), 
                         "login_time":recent_log.get('login_time'), "loggedin_device":recent_log.get('loggedin_device'),
                         "log_id":log_id },
                "status": "already_logged_in",
                "message": "User already logged in from another device"
            }), 409

        # CREATE NEW JWT TOKENS
        additional_claims = {
            "name": user.get("name"),
            "username": user.get("username"),
            "userid": userid,
            "userrole": user.get("role"),
            "mail": user.get("mail"),
            "mobile": user.get("mobile"),
            "indus_circle": user.get("indus_circle"),
            "location": user.get("location"),
        }

        access_token = create_access_token(
            identity=userid,
            additional_claims=additional_claims,
            expires_delta=timedelta(hours=1),
        )

        refresh_token = create_refresh_token(
            identity=userid,
            expires_delta=timedelta(days=30),
        )

        decoded = decode_token(access_token)
        access_jti = decoded["jti"]

        # SAVE NEW SESSION
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO weatherdata.user_sessions (user_id, jti, expires_at)
                VALUES (%s, %s, NOW() + INTERVAL '1 hour')
                ON CONFLICT (user_id)
                DO UPDATE SET
                    jti = EXCLUDED.jti,
                    login_time = NOW(),
                    expires_at = NOW() + INTERVAL '1 hour'
                """,
                (userid, access_jti),
            )

            cur.execute(
                """
                UPDATE weatherdata.licensed_user_auth
                SET online_status = %s,
                    loggedin_device = %s
                WHERE userid = %s
                """,
                ("online", device, userid),
            )

            conn.commit()

        # LOGIN ACTIVITY LOG
        log_id = insert_activity_log(
            userid=userid,
            username=user.get("username"),
            device=device,
            name=user.get("name")
        )
        
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE weatherdata.user_sessions
                SET log_id = %s
                WHERE user_id = %s
                """,
                (log_id, userid),
            ) 
            conn.commit()

        # RESPONSE
        return jsonify({
            "status": "success",
            "message": "Login successful",
            "data": {
                "resultUser": {
                    "userid": userid,
                    "username": user.get("username"),
                    "name": user.get("name"),
                    "userrole": user.get("role"),
                    "indus_circle": user.get("indus_circle"),
                    "location": user.get("location"),
                },
                "logId": log_id,
                "token": access_token,
                "refresh_token": refresh_token,
            }
        }), 200

    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

    finally:
        release_db_conn(conn)
        
@app.route("/force-logout", methods=["POST"])
@cross_origin()
def force_logout_remote():
    conn = None
    try:
        data = request.get_json() or {}
        userId = data.get("userId")
        log_id = data.get("logId")

        if not userId:
            return jsonify({"msg":"userId required"})

        conn = get_db_conn()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Get user_id
            cur.execute(
                """
                SELECT userid 
                FROM weatherdata.licensed_user_auth 
                WHERE userid = %s 
                """,
                (userId,)
            )
            user = cur.fetchone()

        if not user:
            return jsonify({"msg":"User not found"})

        userid = user["userid"]
        
        # Delete all active sessions for this user ---
        try:
            conn = get_db_conn() if conn is None else conn
            with conn.cursor() as cur:
                cur.execute(
                    """
                    DELETE FROM weatherdata.user_sessions
                    WHERE user_id = %s
                    """,
                    (userid,)
                )

                # Update online status
                cur.execute(
                    """
                    UPDATE weatherdata.licensed_user_auth
                    SET online_status = %s
                    WHERE userid = %s
                    """,
                    ("offline", userid)
                )

                conn.commit()
        except Exception as db_err:
            if conn:
                conn.rollback()

        # Update activity log if logId provided ---
        if log_id:
            update_activity_logout(log_id)

        return jsonify({"msg":"User logged out from other device", "status":"success"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

    finally:
        release_db_conn(conn)

@app.route("/user_logout", methods=["POST"])
@cross_origin()
@jwt_required()
def user_logout():
    conn = None
    try:
        identity = get_jwt_identity()
        payload = request.get_json() or {}
        log_id = payload.get("logId")

        # Get current JWT jti
        jwt_data = get_jwt()
        jti = jwt_data.get("jti")

        # Update DB ---
        try:
            conn = get_db_conn() if conn is None else conn
            with conn.cursor() as cur:
                # Mark user offline
                cur.execute(
                    """
                    UPDATE weatherdata.licensed_user_auth
                    SET online_status = %s
                    WHERE userid = %s
                    """,
                    ("offline", identity)
                )

                # Remove session from user_sessions table
                cur.execute(
                    """
                    DELETE FROM weatherdata.user_sessions
                    WHERE user_id = %s AND jti = %s
                    """,
                    (identity, jti)
                )

                conn.commit()
        except Exception as db_err:
            if conn:
                conn.rollback()

        # Update activity log ---
        if log_id:
            update_activity_logout(log_id)

        # Return success response ---
        return jsonify({"msg":"Logout successful", "status":"success"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

    finally:
        release_db_conn(conn)

@app.route("/check-user-session", methods=["POST"])
@cross_origin()
def check_user_session():
    conn = None
    try:
        data = request.get_json() or {}
        username = data.get("username")
        if not username:
            return jsonify({"msg":"username required"})

        conn = get_db_conn()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT 
                    u.userid,
                    u.online_status,
                    u.loggedin_device,
                    l.id AS logId
                FROM weatherdata.licensed_user_auth u
                LEFT JOIN LATERAL (
                    SELECT id
                    FROM weatherdata.weather_user_activity_log
                    WHERE userid = u.userid
                    AND logout_time IS NULL
                    ORDER BY id DESC
                    LIMIT 1
                ) l ON TRUE
                WHERE (u.username = %s OR u.mail = %s)
                """,
                (username, username),
            )
            user = cur.fetchone()

        if not user:
            return jsonify({"msg":"User logged out remotely", "status":"success"}), 200

        device = user.get("loggedin_device", "")
        formatted_device = device or "unknown device"
        is_online = user.get("online_status") == "online"

        return jsonify({
            "msg":"Status fetched",
            "status":"success",
            "data":{
                "userid": user.get("userid"),
                "is_online": is_online,
                "active_device": formatted_device,
                "logId": user.get("logid"),
            },
        }), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/protected", methods=["GET"])
@cross_origin()
@jwt_required()
def protected():
    # simple protected route for testing
    identity = get_jwt_identity()
    return (
        jsonify(
            {
                "status": "success",
                "message": f"You are inside protected route, user={identity}",
            }
        ),
        200,
    )

@app.route("/get-current-weather", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_hourly_data():
    conn = get_db_conn()
    try:
        data = request.get_json()

        if not data:
            return jsonify({"error": "No JSON data provided"}), 400

        selected_date = data.get("params")["selectedDate"]
        
        with conn.cursor() as cursor:
            query = f"""
                SELECT temp_c, chance_of_rain, wind_kph, humidity, vis_km, latitude, longitude, city_name
                    FROM weatherdata.weather_hourly_data_all_india
                    WHERE time = '{selected_date}'
                ORDER BY city_name ASC;
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]

            return jsonify({"status": "success", "data": result})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get_circle_weather_min_max", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_circle_weather_min_max():
    conn = get_db_conn()
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No JSON data provided"}), 400

        circle = data.get("circle")
        if not circle:
            return (
                jsonify({"status": "error", "message": "Missing 'circle' in request"}),
                400,
            )
            
        with conn.cursor() as cursor:
            query = """
                SELECT 
                    MIN(temp_min) AS temp_min,
                    MAX(temp_max) AS temp_max,
                    MIN(wind) AS wind_min,
                    MAX(wind) AS wind_max,
                    MIN(rain_precip) AS rain_min,
                    MAX(rain_precip) AS rain_max,
                    MIN(humidity) AS humidity_min,
                    MAX(humidity) AS humidity_max,
                    MIN(visibility) AS visibility_min,
                    MAX(visibility) AS visibility_max
                FROM weatherdata.district_wise_7dayfc_severity
                WHERE indus_circle = %s AND days = 'day1';
            """
            cursor.execute(query, (circle,))
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "circle": circle, "data": result})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally: 
        release_db_conn(conn)

@app.route("/get-earthquake", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_earthquake_data():
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """SELECT * FROM weatherdata.earthquakes WHERE "time" >= NOW() - INTERVAL '2 days';"""
            cursor.execute(query)
            rows = cursor.fetchall()
            # Optional: Get column names
            colnames = [desc[0] for desc in cursor.description]
            # Convert to list of dicts
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get-today-disasters", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_today_disasters():
    conn = get_db_conn()
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data provided"}), 400
        hazard_type = data.get("params")["hazardType"]
        severity_type = data.get("params")["severityType"]

        with conn.cursor() as cursor:
            query = f"""SELECT sender, TO_CHAR(sent, 'DD-MM-YYYY HH24:MI') as sent, event, severity, certainty, TO_CHAR(effective, 'DD-MM-YYYY HH24:MI') as effective, TO_CHAR(onset, 'DD-MM-YYYY HH24:MI') as onset, TO_CHAR(expires, 'DD-MM-YYYY HH24:MI') as expires, headline, description,id,
                        "areaDesc", geocode_name_0 as state, st_asgeojson(geom) as geometry 
                    FROM weatherdata.disaster_ndma WHERE sent::date = CURRENT_DATE AND ('{hazard_type}' = 'All' OR event LIKE '%{hazard_type}%') AND ('{severity_type}' = 'All' OR severity LIKE '%{severity_type}%') order by sent desc;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]

            return jsonify({"status": "success", "data": result})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get-selected-disasters", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_selected_disasters():
    conn = get_db_conn()
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data provided"}), 400
        id = data.get("id")
        # SQL to get current hour data
        with conn.cursor() as cursor:
            query = f"""SELECT sender, TO_CHAR(sent, 'DD-MM-YYYY HH24:MI') as sent, event, severity, certainty, TO_CHAR(effective, 'DD-MM-YYYY HH24:MI') as effective, TO_CHAR(onset, 'DD-MM-YYYY HH24:MI') as onset, TO_CHAR(expires, 'DD-MM-YYYY HH24:MI') as expires, headline, description,id,
                        "areaDesc", geocode_name_0 as state, st_asgeojson(geom) as geometry 
                    FROM weatherdata.disaster_ndma WHERE sent::date = CURRENT_DATE AND id = {id} order by sent desc;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

# Hazards (Disaster) Dropdown and Data methods
@app.route("/get-hazards-list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_hazards_list():
    conn =  get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """SELECT DISTINCT 
                        TRIM(split_part(val, 'and', 1)) AS hazard_list
                        FROM (
                            SELECT sent, unnest(string_to_array(event, ',')) AS val
                            FROM weatherdata.disaster_ndma WHERE sent::date = CURRENT_DATE 
                        ) AS sub order by hazard_list asc ;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get-severity-list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_severity_list():
    conn =  get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """SELECT DISTINCT 
                        TRIM(split_part(val, 'and', 1)) AS severity
                    FROM (
                        SELECT sent, unnest(string_to_array(severity, ',')) AS val
                        FROM weatherdata.disaster_ndma WHERE sent::date = CURRENT_DATE 
                    ) AS sub order by severity asc ;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

# Sending Selected Tower Report to users
@app.route("/api-send-report", methods=["POST"])
@cross_origin("*")
@jwt_required()
def send_report():
    file = request.files["file"]
    save_path = os.path.join("output", file.filename)
    file.save(save_path)
    user_mail = request.form.get("userMail")
    RECIVERS = user_mail.split(",")
    # Format current date & time
    current_date = datetime.now().strftime("%d %b %Y")  # e.g. "22 Aug 2025"
    current_time = datetime.now().strftime("%H:00")  # e.g. "10:00"

    yag = yagmail.SMTP(user="post@mlinfomap.com", password="tmisxyakmbllotlw")
    yag.send(
        to=RECIVERS,
        # [
        #     "subodh@mlinfomap.com",
        #     "ravikumar7277rv@gmail.com",
        #     "Aditya@mlinfomap.com",
        #     "rizwansiddiqui5225@gmail.com",
        # ],
        subject=f"Bad Weather Alert – {current_date}, {current_time} Hrs",
        contents=[
            "Please find attached the tower weather alert.",
            yagmail.inline(save_path),
            save_path,
        ],
    )
    return {"status": "sent"}

@app.route("/weather_user_activity", methods=["POST"])
@cross_origin("*")
@jwt_required()
def user_activity():
    conn =  get_db_conn()
    try:
        payload = request.json
        activity_type = payload.get("type")
        data = payload.get("data", {})
        cur = conn.cursor(cursor_factory=DictCursor)
        if activity_type == "login":
            # Insert new row
            columns = ", ".join(data.keys())
            values_placeholders = ", ".join(["%s"] * len(data))
            values = list(data.values())

            query = f"""
                INSERT INTO weatherdata.weather_user_activity_log ({columns})
                VALUES ({values_placeholders})
                RETURNING id;
            """
            cur.execute(query, values)
            new_id = cur.fetchone()["id"]
            conn.commit()
            return jsonify(
                {
                    "status": "success",
                    "message": "Log Inserted Successfully",
                    "id": new_id,
                }
            )

        elif activity_type == "update":
            row_id = payload.get("id")
            if not row_id:
                return (
                    jsonify({"status": "error", "message": "Missing id for update"}),
                    400,
                )

            set_clause = ", ".join([f"{col} = %s" for col in data.keys()])
            values = list(data.values()) + [row_id]

            query = f"""
                UPDATE weatherdata.weather_user_activity_log
                SET {set_clause}
                WHERE id = %s;
            """
            cur.execute(query, values)
            conn.commit()
            return jsonify(
                {"status": "success", "message": "Row updated", "id": row_id}
            )

        else:
            return jsonify({"status": "error", "message": "Invalid type"}), 400
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

# fetching the all users from licensed_user_auth table
@app.route("/get-user-list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_user_list():
    conn =  get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """
                        select id, userid, "name",username, status, "role", mail, mobile, indus_circle, status_activation_date, status_deactivation_date 
                        from weatherdata.licensed_user_auth where role not in ('H_MGMT') order by userid asc;
                        """
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get_circle_list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_circle_list():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        indus_circle = payload.get("circle")  
        cursor = conn.cursor(cursor_factory=DictCursor)
        query = f"""select indus_circle, indus_circle_name, xx as longitude, yy as latitude from weatherdata.indus_circle_geomerty
                    where ( 'All Circle' = '{indus_circle}' or indus_circle = '{indus_circle}') order by indus_circle;                                  
                """
        cursor.execute(query)
        result = cursor.fetchall()
        circle_list = []
        all_circle = {}
        for row in result:
            if row['indus_circle'] == 'M&G':
                all_circle = {
                    "label": "All Circle",
                    "value": f"{row['latitude']},{row['longitude']}",
                    "full_name": row["indus_circle_name"]
                }
            circle_list.append({
                "label": row["indus_circle"],
                "value": f"{row['latitude']},{row['longitude']}",
                "full_name": row["indus_circle_name"]
            })
             
        # circle_list = sorted(circle_list, key=lambda x: x["label"])
        if all_circle:
            circle_list.insert(0, all_circle)
        return jsonify({"status": "success", "data": circle_list})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)
    
@app.route("/get_district_list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_district_list():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        circle = payload.get("circle")
        with conn.cursor() as cursor:
            query = """ select distinct district from weatherdata.district_geometry 
                    where ('All Circle' = %s or indus_circle = %s ) and indus_circle is not null and indus_circle <> '' order by district ASC;                                  
                    """
            cursor.execute(query, (circle,circle))
            result = cursor.fetchall()
            district_list = [{"district": row[0]} for row in result]
            return jsonify({"status": "success", "data": district_list})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)
    
@app.route("/inserted_hazard_circle_list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def inserted_hazard_circle_list():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        hazard_type = payload.get("hazard")
        table_map = {
            "Flood": "hazard_flood",
            "Cyclone": "hazard_cyclone",
            "Snowfall": "hazard_snowfall",
            "Avalanche": "hazard_avalanche",
            "Cloudburst": "hazard_cloudburst",
            "Lightning": "hazard_lightning",
            "Landslide": "hazard_landslide"
            }
        table_name = table_map.get(hazard_type, None)
        with conn.cursor() as cursor:
            query = f"""
                SELECT DISTINCT indus_circle
                FROM weatherdata.{table_name}
                WHERE insert_at::date = CURRENT_DATE;
            """
            cursor.execute(query)
            result = cursor.fetchall()
            indus_circle_list = [
                {"indus_circle": row[0]} 
                for row in result if row[0] is not None
            ]
            return jsonify({
                "status": "success",
                "data": indus_circle_list
            }), 200

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)
    
@app.route("/insert-hazards", methods=["POST"])
@cross_origin("*")
@jwt_required()
def insert_hazards_forecast():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        items = payload.get("data", [])
        hazard_type = payload.get("hazard")
        table_map = {
            "Flood": "hazard_flood",
            "Cyclone": "hazard_cyclone",
            "Snowfall": "hazard_snowfall",
            "Avalanche": "hazard_avalanche",
            "Cloudburst": "hazard_cloudburst",
            "Lightning": "hazard_lightning",
            "Landslide": "hazard_landslide"
            }
        table_name = table_map.get(hazard_type, None)
        expanded_items = format_hazard_records(items)
        with conn.cursor() as cursor:
            insert_query = f"""
                INSERT INTO weatherdata.{table_name} 
                (days, "date", indus_circle, district, hazard_value, description, severity)
                VALUES %s
            """
            rows_to_insert = [
                (
                    x["day"],
                    x["date"],
                    x["circle"],
                    x["district"],
                    x["hazardValue"],
                    x["description"],
                    x["severity"]
                )
                for x in expanded_items
            ]
            execute_values(cursor, insert_query, rows_to_insert)
            conn.commit()
            return jsonify({
                "status": "success",
                "message": f"{len(rows_to_insert)} Records insert successfully. " ,
            })

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)
  
############### Insert & Update ###############  
# @app.route("/insert-hazards", methods=["POST"])
# @cross_origin("*")
# @jwt_required()
# def insert_hazards_forecast():
#     conn =  get_db_conn()
#     try:
#         payload = request.get_json()
#         items = payload.get("data", [])
#         hazard_type = payload.get("hazard")
        
        # table_map = {
        #     "Flood": "hazard_flood",
        #     "Cyclone": "hazard_cyclone",
        #     "Snowfall": "hazard_snowfall",
        #     "Avalanche": "hazard_avalanche",
        #     "Cloudburst": "hazard_cloudburst",
            # "Lightning": "hazard_lightning",
            # "Landslide": "hazard_landslide"
        #     }
        # table_name = table_map.get(hazard_type, None)

#         expanded_items = format_hazard_records(items)

#         cursor = conn.cursor()

#         from psycopg2.extras import execute_values

#         insert_query = f"""
#             INSERT INTO weatherdata.{table_name}
#             (days, "date", indus_circle, district, hazard_value, description, severity)
#             VALUES %s
#             ON CONFLICT ("date", district, indus_circle)
#             DO UPDATE SET
#                 days = EXCLUDED.days,
#                 hazard_value = EXCLUDED.hazard_value,
#                 description = EXCLUDED.description,
#                 severity = EXCLUDED.severity;
#         """

#         rows_to_insert = [
#             (
#                 x["day"],
#                 x["date"],
#                 x["circle"],
#                 x["district"],
#                 x["hazardValue"],
#                 x["description"],
#                 x["severity"]
#             )
#             for x in expanded_items
#         ]

#         execute_values(cursor, insert_query, rows_to_insert)
#         conn.commit()
#         return jsonify({
#             "status": "success",
#             "message": f"{len(rows_to_insert)} Records inserted/updated successfully."
#         })

#     except Exception as e:
#         return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
#     finally:
#         release_db_conn(conn)
    
@app.route("/get-hazards", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_hazards_forecast():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        hazard_type = payload.get("hazard")
        table_map = {
            "Flood": "hazard_flood",
            "Cyclone": "hazard_cyclone",
            "Snowfall": "hazard_snowfall",
            "Avalanche": "hazard_avalanche",
            "Cloudburst": "hazard_cloudburst",
            "Lightning": "hazard_lightning",
            "Landslide": "hazard_landslide"
            }
        table_name = table_map.get(hazard_type, None)
        with conn.cursor() as cursor:
            query = f"""SELECT * FROM weatherdata.{table_name} hf WHERE hf.date >= CURRENT_DATE;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = []
            for row in rows:
                item = dict(zip(colnames, row))
                if "district" in item and isinstance(item["district"], str):
                    item["district"] = [d.strip() for d in item["district"].split(",")]
                result.append(item)
            return jsonify({
                "status": "success",
                "data": result
            })
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get-district-wise-hazards", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_district_wise_hazards_forecast():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        hazard_type = payload.get("hazardType")
        circle = payload.get("circle") 
        table_map = {
        "Flood": "hazard_flood",
        "Cyclone": "hazard_cyclone",
        "Snowfall": "hazard_snowfall",
        "Avalanche": "hazard_avalanche",
        "Cloudburst": "hazard_cloudburst",
        "Lightning": "hazard_lightning",
        "Landslide": "hazard_landslide"
        }
        table_name = table_map.get(hazard_type, None)
        
        sql = f""" 
           WITH days AS (
                SELECT 'Day1' AS day UNION ALL
                SELECT 'Day2' UNION ALL
                SELECT 'Day3' UNION ALL
                SELECT 'Day4' UNION ALL
                SELECT 'Day5' UNION ALL
                SELECT 'Day6' UNION ALL
                SELECT 'Day7'
            ),

            exploded AS (
                SELECT
                    hs.indus_circle,
                    hs.days,
                    hs.date,
                    TRIM(district_item) AS district,
                    hs.severity,
                    CASE hs.severity
                        WHEN 'Extreme'  THEN 1
                        WHEN 'High'     THEN 2
                        WHEN 'Moderate' THEN 3
                        WHEN 'Low'      THEN 4
                        ELSE 5
                    END AS sev_rank
                FROM weatherdata.{table_name} hs
                CROSS JOIN LATERAL (
                    SELECT unnest(string_to_array(hs.district, ',')) AS district_item
                ) u
                WHERE DATE(hs.insert_at) = CURRENT_DATE - 1
                AND hs.indus_circle = '{circle}'
                AND hs.district IS NOT NULL
                AND hs.district <> ''
            ),

            best_per_district AS (
                SELECT indus_circle, days, district, date, severity
                FROM (
                    SELECT
                        indus_circle,
                        days,
                        district,
                        date,
                        severity,
                        sev_rank,
                        ROW_NUMBER() OVER (PARTITION BY indus_circle, days, district ORDER BY sev_rank) AS rn
                    FROM exploded
                ) t
                WHERE rn = 1
            )

            SELECT
                a.district,
                a.indus_circle,
                d.day AS days,
                b.date,
                COALESCE(b.severity, 'Other') AS severity
            FROM weatherdata.district_geometry a
            CROSS JOIN days d
            LEFT JOIN best_per_district b
                ON b.indus_circle = a.indus_circle
            AND b.days = d.day
            AND b.district = a.district
            WHERE a.district <> 'Data Not Available'
            AND a.indus_circle = '{circle}'
            ORDER BY a.district, d.day;
        """      
        df = pd.read_sql_query(sql, conn)

        result = []
        for district, group in df.groupby("district"):
            district_data = {"district": district}

            for _, row in group.iterrows():
                district_data[row["days"]] = {
                    "date": row["date"],
                    "severity": "No Risk" if row["severity"] == "Other" or row["severity"] == "Low" else f"{row['severity']} Risk",
                    "for_color":row["severity"],
                    "indus_circle": row["indus_circle"]
                }
            result.append(district_data)

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get-hazard-affected-district", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_hazard_affected_districts():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        # hazard_type = payload.get("hazardType")
        circle = payload.get("circle")
        table_map = {
            "Flood": "hazard_flood",
            "Cyclone": "hazard_cyclone",
            "Snowfall": "hazard_snowfall",
            "Avalanche": "hazard_avalanche",
            "Cloudburst": "hazard_cloudburst",
            "Lightning": "hazard_lightning",
            "Landslide": "hazard_landslide"
        }

        sql = f"""
                select days, date, indus_circle, district, severity, 'Avalanche' as hazard from weatherdata.hazard_avalanche
                where DATE(insert_at) = CURRENT_DATE - 1 and indus_circle = '{circle}' and days = 'Day1'
                union 
                select days, date, indus_circle, district, severity, 'Cloudburst' as hazard from weatherdata.hazard_cloudburst
                where DATE(insert_at) = CURRENT_DATE - 1 and indus_circle = '{circle}' and days = 'Day1'
                union
                select days, date, indus_circle, district, severity, 'Cyclone' as hazard from weatherdata.hazard_cyclone
                where DATE(insert_at) = CURRENT_DATE - 1 and indus_circle = '{circle}' and days = 'Day1'
                union  
                select days, date, indus_circle, district, severity, 'Flood' as hazard from weatherdata.hazard_flood
                where DATE(insert_at) = CURRENT_DATE - 1 and indus_circle = '{circle}' and days = 'Day1'
                union
                select days, date, indus_circle, district, severity, 'Lightning' as hazard from weatherdata.hazard_lightning
                where DATE(insert_at) = CURRENT_DATE - 1 and indus_circle = '{circle}' and days = 'Day1'
                union
                select days, date, indus_circle, district, severity, 'Snowfall' as hazard from weatherdata.hazard_snowfall
                where DATE(insert_at) = CURRENT_DATE - 1 and indus_circle = '{circle}' and days = 'Day1';
        """

        df = pd.read_sql_query(sql, conn)
        severity_levels = ["Extreme", "High", "Moderate", "Low"]
        # Initialize empty structure for all hazards
        hazard_dict = {
            "Avalanche": {sev: [] for sev in severity_levels},
            "Cloudburst": {sev: [] for sev in severity_levels},
            "Cyclone": {sev: [] for sev in severity_levels},
            "Flood": {sev: [] for sev in severity_levels},
            "Lightning": {sev: [] for sev in severity_levels},
            "Snowfall": {sev: [] for sev in severity_levels},
        }
        if df.empty:
            return jsonify({"status": "success", "data": hazard_dict})
        for _, row in df.iterrows():
            hazard = row.get("hazard")
            sev = row.get("severity")

            if hazard not in hazard_dict:
                continue
            if sev not in severity_levels:
                continue

            districts = (
                row["district"].split(",")
                if isinstance(row["district"], str) and row["district"].strip()
                else []
            )

            hazard_dict[hazard][sev] = districts
        return jsonify({"status": "success", "data": [hazard_dict]})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/update_user_status", methods=["POST"])
@cross_origin("*")
@jwt_required()
def update_user_status():
    conn =  get_db_conn()
    try:
        payload = request.json
        userid = payload.get("id")
        status = payload.get("status")

        if not userid:
            return jsonify({"status": "error", "message": "Missing id for update"}), 400
        with conn.cursor() as cur:
            current_datetime = datetime.now().strftime("%d %b %Y %H:%M")
            if status == "active":
                count_query = """
                    SELECT COUNT(*)
                    FROM weatherdata.licensed_user_auth
                    WHERE status = 'active' AND userid != 'ml.admin'
                """
                cur.execute(count_query)
                active_count = cur.fetchone()[0]

                cur.execute("SELECT allowed_users FROM weatherdata.weather_user_license")
                allowed_users_count = cur.fetchone()[0]

                if active_count >= allowed_users_count:
                    cur.close()
                    conn.close()
                    return (
                        jsonify(
                            {
                                "status": "Bad Request",
                                "message": "Active users cannot exceed the limit",
                            }
                        ),
                        400,
                    )
                else:
                    query = """
                                UPDATE weatherdata.licensed_user_auth
                                SET status = %s, status_activation_date = %s, status_deactivation_date = NULL
                                WHERE userid = %s
                            """
                    params = (status, current_datetime, userid)
                    cur.execute(query, params)
                    conn.commit()
                    return jsonify(
                        {"status": "success", "message": "User status updated successfully"}
                    )

            else:
                query = """
                    UPDATE weatherdata.licensed_user_auth
                    SET status = %s, status_deactivation_date = %s
                    WHERE userid = %s
                """
                params = (status, current_datetime, userid)
                cur.execute(query, params)
                conn.commit()

                cur.close()
                conn.close()

                return jsonify(
                    {"status": "success", "message": "User status updated successfully"}
                )

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        conn =  get_db_conn()

@app.route("/update_user", methods=["POST"])
@cross_origin("*")
@jwt_required()
def update_user():
    conn =  get_db_conn()
    try:
        payload = request.json
        data = payload.get("data")
        userid = payload.get("id")
        if not userid:
            return (
                jsonify({"status": "error", "message": "Missing id for update"}),
                400,
            )

        set_clause = ", ".join([f"{col} = %s" for col in data.keys()])
        values = list(data.values()) + [userid]
        with conn.cursor() as cursor:
            query = f"""
                    UPDATE weatherdata.licensed_user_auth
                    SET {set_clause}
                    WHERE userid = %s;
                """
            cursor.execute(query, values)
            conn.commit()
            return jsonify({"status": "success", "message": "User updated Successfully"})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/add_new_user", methods=["POST"])
@cross_origin("*")
@jwt_required()
def add_new_user():
    conn =  get_db_conn()
    try:
        payload = request.json
        data = payload.get("data")
        if not data:
            return jsonify({"status": "error", "message": "User data is required"}), 400
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500

        status = data.get("status")
        mail = data.get("mail")
        username = data.get("username")

        # CHECK DUPLICATE USER
        with conn.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(
                "SELECT * FROM weatherdata.licensed_user_auth WHERE username = %s OR mail = %s",
                (username, mail),
            )
            user = cur.fetchone()
            if user:
                return jsonify({"status": "Bad Request", "message": "User already exists"}), 400

        # REQUIRED FIELDS
        required_fields = [
            "name", "username", "password", "status",
            "role", "mail", "mobile", "indus_circle"
        ]
        missing_fields = [f for f in required_fields if not data.get(f)]

        if missing_fields:
            return jsonify({
                "status": "error",
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400

        # ACTIVE USER LIMIT CHECK
        if status == "active":
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM weatherdata.licensed_user_auth WHERE status = 'active'")
                active_count = cur.fetchone()[0]

                cur.execute("SELECT allowed_users FROM weatherdata.weather_user_license")
                allowed_users_count = cur.fetchone()[0]

                if active_count >= allowed_users_count:
                    return jsonify({
                        "status": "Bad Request",
                        "message": "Active users cannot exceed the licensed limit"
                    }), 400

        # HANDLE STATUS DATES
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if status == "active":
            data["status_activation_date"] = now
            data["status_deactivation_date"] = None
        else:
            data["status_activation_date"] = None
            data["status_deactivation_date"] = None

        # INSERT USER
        columns = ", ".join(data.keys())
        placeholders = ", ".join(["%s"] * len(data))
        values = list(data.values())
        insert_query = f"""
            INSERT INTO weatherdata.licensed_user_auth ({columns})
            VALUES ({placeholders})
        """
        with conn.cursor() as cur:
            cur.execute(insert_query, values)
            conn.commit()
        return jsonify({"status": "success", "message": "User added successfully"}), 201
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get-user-license", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_user_license():
    conn =  get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """SELECT * FROM weatherdata.weather_user_license order by id asc;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/update-user-license", methods=["POST"])
@cross_origin("*")
@jwt_required()
def update_user_license():
    conn =  get_db_conn()
    try:
        payload = request.json
        allowed_users = int(payload.get("allowed_users"))
        with conn.cursor() as cursor:
            query = f"""UPDATE weatherdata.weather_user_license SET allowed_users = {allowed_users};"""
            cursor.execute(query)
            conn.commit()
            return (
                jsonify(
                    {"status": "success", "message": "User License updated successfully"}
                ),
                200,
            )
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/change-user-password", methods=["POST"])
@cross_origin("*")
@jwt_required()
def change_user_password():
    conn =  get_db_conn()
    try:
        payload = request.json
        mail = payload.get("mail")
        password = payload.get("password")
        with conn.cursor() as cursor:
            query = f"""UPDATE gis_admin.user_auth_mlfinfo_drawing SET password = '{password}' WHERE mail = '{mail}';"""
            cursor.execute(query)
            conn.commit()
            return (
                jsonify(
                    {"status": "success", "message": "User password updated successfully"}
                ),
                200,
            )

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/change-password", methods=["POST"])
@cross_origin("*")
@jwt_required()
def change_password():
    conn =  get_db_conn()
    try:
        payload = request.json
        user_id = payload.get("userId")
        old_password = payload.get("oldPassword")
        new_password = payload.get("newPassword")

        if not user_id or not old_password or not new_password:
            return jsonify({"status": "Bad Request", "message": "Missing fields"}), 400
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT password FROM gis_admin.user_auth_mlfinfo_drawing WHERE userid = %s",
                (user_id,),
            )
            row = cursor.fetchone()
            if not row:
                return jsonify({"status": "Bad Request", "message": "User not found"}), 400

        stored_password = row[0]
        if stored_password != old_password:
            return (
                jsonify(
                    {"status": "Bad Request", "message": "Old password doesn't match"}
                ),
                400,
            )
        cur.execute(
            "UPDATE gis_admin.user_auth_mlfinfo_drawing SET password = %s WHERE userid = %s",
            (new_password, user_id),
        )
        conn.commit()
        return (
            jsonify(
                {"status": "success", "message": "User password updated successfully"}
            ),
            200,
        )
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

    finally:
        release_db_conn(conn)

@app.route("/fetch-kpi-range", methods=["POST"])
@cross_origin("*")
@jwt_required()
def fetch_kpi_range():
    conn =  get_db_conn()
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        query = "SELECT * FROM weatherdata.weather_kpi_controls WHERE indus_circle IS NOT NULL;"
        cur.execute(query)
        rows = cur.fetchall()
        if not rows:
            return jsonify({"status": "success", "data": []}), 200
        grouped_data = {}
        for row in rows:
            circle = row["indus_circle"]
            if circle not in grouped_data:
                grouped_data[circle] = []
            grouped_data[circle].append(row)

        return jsonify({"status": "success", "data": grouped_data}), 200
    except Exception as e:
        print(e)
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/update-kpi-range", methods=["POST"])
@cross_origin("*")
@jwt_required()
def update_kpi_range():
    conn =  get_db_conn()
    try:
        data = request.get_json() 
        circle = data.get("circle")
        if not circle:
            return jsonify({"status": "error", "message": "Circle is required"}), 400
        # Remove 'circle' since it's not an actual DB column
        update_data = {k: v for k, v in data.items() if k not in ["circle"]}
        if not update_data:
            return jsonify({"status": "error", "message": "No fields to update"}), 400
        cur = conn.cursor()
        color_updates = {k: v for k, v in update_data.items() if "color" in k.lower()}
        normal_updates = {
            k: v for k, v in update_data.items() if "color" not in k.lower()
        }
        for key, value in color_updates.items():
            query = f"""
                UPDATE weatherdata.weather_kpi_controls
                SET {key} = %s WHERE indus_circle = %s
            """
            cur.execute(query, (value, circle))
        if normal_updates:
            set_clause = ", ".join([f"{key} = %s" for key in normal_updates.keys()])
            values = list(normal_updates.values())
            values.append(circle)  
            query = f"""
                UPDATE weatherdata.weather_kpi_controls
                SET {set_clause}
                WHERE indus_circle = %s
            """
            cur.execute(query, values)
        conn.commit()
        return jsonify(
            {
                "status": "success",
                "message": f"KPI range updated successfully.",
            }
        )
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)
        
@app.route("/get_log_min_max_date", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_log_min_max_date():
    conn =  get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """ select min(DATE(login_time)) as min_date,  max(DATE(login_time)) as max_date 
                   from weatherdata.weather_user_activity_log;                                  
                    """
            cursor.execute(query)
            row = cursor.fetchone()
            columns = [desc[0] for desc in cursor.description]
            record = dict(zip(columns, row))
            return jsonify({"status": "success", "data": record})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)
        
@app.route("/get_log_summary_date_wise", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_log_summary_date_wise():
    conn =  get_db_conn()
    try:
        data = request.get_json() 
        start_date = data.get("startDate")
        end_date = data.get("endDate")
        condition = f"WHERE login_time >= '{start_date}' AND login_time < DATE '{end_date}' + INTERVAL '1 day' " if start_date != '' and end_date != '' else ''
        with conn.cursor() as cursor:
            query = f""" SELECT
                        DATE(login_time) AS login_date,
                        name,
                        userid,
                        COUNT(*) AS login_count,
                        (
                        floor(extract(epoch FROM SUM(logout_time - login_time)) / 3600)::int
                        || ':' ||
                        lpad(floor(mod(extract(epoch FROM SUM(logout_time - login_time)), 3600) / 60)::int::text, 2, '0')
                        || ':' ||
                        lpad(floor(mod(extract(epoch FROM SUM(logout_time - login_time)), 60))::int::text, 2, '0')
                        ) AS duration
                    FROM weatherdata.weather_user_activity_log
                    {condition}
                    GROUP BY DATE(login_time), name, userid
                    ORDER BY login_date DESC, name;                                  
                    """
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/fetch_dashboad_usage", methods=["POST"])
@cross_origin("*")
@jwt_required()
def fetch_dashboard_usages():
    conn =  get_db_conn()
    try:
        data = request.get_json()
        log_date = data.get("logDate")
        user_id = data.get("userid")

        with conn.cursor() as cursor:
            query = f"""select * from weatherdata.weather_user_activity_log 
                    where userid = '{user_id}' and DATE(login_time) = '{log_date}'
                    order by login_time DESC;"""
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            ACTION_COLUMNS = [
            "today_btn_clicked",
            "tomorrow_btn_clicked",
            "today_temp_clicked",
            "today_rain_clicked",
            "today_wind_clicked",
            "today_humidity_clicked",
            "today_visibility_clicked",
            "tomorrow_temp_clicked",
            "tomorrow_rain_clicked",
            "tomorrow_wind_clicked",
            "tomorrow_humidity_clicked",
            "tomorrow_visibility_clicked",
            "tower_clicked",
            "lasso_tool_clicked",
            "alert_send",
            "alert_send_time",
            "alert_send_user",
            "search_term",
            "search_time",
            "hazard_type_selected",
            "severity_selected",
            "view_on_map_clicked",
            "dashboard_clicked",
            "circlelevel_clicked",
            "pandindia_clicked",
            "usage_clicked",
            "thvscore_clicked",
            "dashboard_hourly_weather_clicked",
            "dashboard_seven_day_forecast_clicked",
            "dashboard_hazard_alert_clicked",
            "dashboard_hazard_type_clicked",
            "dashboard_hazard_severity_clicked",
            "dashboard_view_map_clicked",
            "circle_pdf_download",
            "circle_level_clicked",
            "circle_weather_param_breakdown_view",
            "circle_today_risk_weather_view",
            "circle_today_risk_hazard_view",
            "circle_weather_forecast_view",
            "circle_hazard_forecast_view",
            "cyclone_clicked",
            "cyclone_map_layer_checked_unchecked",
            "cyclone_severity_table_export",
            "circle_weather_forecast_rainfall",
            "circle_weather_forecast_accu_rainfall",
            "circle_weather_forecast_wind",
            "circle_weather_forecast_humidity",
            "circle_weather_forecast_visibility",
            "circle_weather_forecast_temperature",
            "circle_weather_hazard_cyclone",
            "circle_weather_hazard_lightning",
            "circle_weather_hazard_flood",
            "circle_weather_hazard_snowfall",
            "circle_weather_hazard_avalanche"
            ]
            ACTION_COLUMNS = set(ACTION_COLUMNS)
            final_result = []
            for row in rows:
                record = dict(zip(columns, row))

                actions = []

                for col in ACTION_COLUMNS:
                    val = record.get(col)

                    if val == 'true':
                        actions.append(col)
                    elif val not in (False, None, ""):
                        actions.append(f"{col}:{val}")

                # attach action
                record["action"] = actions

                # remove action columns from top-level
                for col in ACTION_COLUMNS:
                    record.pop(col, None)

                final_result.append(record)
                
            return (
                jsonify(
                    {
                        "status": "success",
                        "message": "Dashboard usage fetched successfully.",
                        "data": final_result,
                    }
                ),
                200,
            )
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/fetch_circle_report", methods=["POST"])
@cross_origin("*")
def circle_report_data():
    conn =  get_db_conn()
    try:
        data = request.get_json()
        circle = data.get("circle")
        with conn.cursor() as cursor:
            sql = f""" 
            SELECT 
                days,
                MIN(date) AS date,
                severity_type,
                STRING_AGG(DISTINCT district, ', ') AS districts,
                SUM(CASE WHEN severity = 'Extreme' THEN 1 ELSE 0 END) AS extreme,
                SUM(CASE WHEN severity = 'High' THEN 1 ELSE 0 END) AS high,
                SUM(CASE WHEN severity = 'Moderate' THEN 1 ELSE 0 END) AS moderate,
                SUM(CASE WHEN severity = 'Other' THEN 1 ELSE 0 END) AS other
            FROM (
                SELECT days, date, district, 'Temperature_Max' AS severity_type, temp_max_severity AS severity 
                FROM weatherdata.district_wise_7dayfc_severity WHERE ('All Circle' = %s or indus_circle = %s)

                UNION ALL

                SELECT days, date, district, 'Temperature_Min' AS severity_type, temp_min_severity AS severity 
                FROM weatherdata.district_wise_7dayfc_severity WHERE ('All Circle' = %s or indus_circle = %s)
                
                UNION ALL
                
                SELECT days, date, district, 'Rainfall', rain_severity 
                FROM weatherdata.district_wise_7dayfc_severity WHERE ('All Circle' = %s or indus_circle = %s)
                
                UNION ALL
                
                SELECT days, date, district, 'Wind', wind_severity 
                FROM weatherdata.district_wise_7dayfc_severity WHERE ('All Circle' = %s or indus_circle = %s)
                
                UNION ALL
                
                SELECT days, date, district, 'Humidity', humidity_severity 
                FROM weatherdata.district_wise_7dayfc_severity WHERE ('All Circle' = %s or indus_circle = %s)
                
                UNION ALL
                
                SELECT days, date, district, 'Visibility', visibility_severity 
                FROM weatherdata.district_wise_7dayfc_severity WHERE ('All Circle' = %s or indus_circle = %s)
                
            ) AS severity_data
            GROUP BY days, severity_type
            ORDER BY days, severity_type;
            """

            df = pd.read_sql_query(sql, conn, params=[circle] * 12)
            merged = {}
            for idx, row in df.iterrows():
                day = row["days"]
                severity = row["severity_type"]
                if day not in merged:
                    merged[day] = {}
                merged[day][severity] = {
                    "date": row["date"],
                    "districts": row["districts"],
                    "extreme": row["extreme"],
                    "high": row["high"],
                    "moderate": row["moderate"],
                    "other": row["other"],
                }
            severity_color = fetch_severity_colors(circle)
            return merged, severity_color[0]
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/fetch_district_names_severity_wise", methods=["POST"])
@cross_origin("*")
@jwt_required()
def fetch_district_names_severity_wise_7days():
    conn =  get_db_conn()
    try:
        data = request.get_json()  
        circle = data.get("circle")
        sql = f""" 
            SELECT 
                    days,
                    "date",
                    severity_type,
                    STRING_AGG(CASE WHEN severity = 'Extreme' THEN district END, ', ') AS extreme_districts,
                    STRING_AGG(CASE WHEN severity = 'High' THEN district END, ', ') AS high_districts,
                    STRING_AGG(CASE WHEN severity = 'Moderate' THEN district END, ', ') AS moderate_districts,
                    STRING_AGG(CASE WHEN severity = 'Low' THEN district END, ', ') AS low_districts
                FROM (
                    SELECT days,"date", district, 'Temperature_Max' AS severity_type, temp_max_severity AS severity 
                    FROM weatherdata.district_wise_7dayfc_severity 
                    WHERE indus_circle = '{circle}'
                    
                    UNION ALL

                    SELECT days,"date", district, 'Temperature_Min' AS severity_type, temp_min_severity AS severity 
                    FROM weatherdata.district_wise_7dayfc_severity 
                    WHERE indus_circle = '{circle}'
                    
                    UNION ALL
                    
                    SELECT days,"date", district, 'Rainfall', rain_severity 
                    FROM weatherdata.district_wise_7dayfc_severity 
                    WHERE indus_circle = '{circle}'
                    
                    UNION ALL
                    
                    SELECT days,"date", district, 'Wind', wind_severity 
                    FROM weatherdata.district_wise_7dayfc_severity 
                    WHERE indus_circle = '{circle}'
                    
                    UNION ALL
                    
                    SELECT days,"date", district, 'Visibility', visibility_severity 
                    FROM weatherdata.district_wise_7dayfc_severity 
                    WHERE indus_circle = '{circle}'
                    
                    UNION ALL
                    
                    SELECT days,"date", district, 'Humidity', humidity_severity 
                    FROM weatherdata.district_wise_7dayfc_severity 
                    WHERE indus_circle = '{circle}'
                ) AS severity_data
                GROUP BY days, severity_type,"date"
                ORDER BY days, severity_type;
            """

        df = pd.read_sql_query(sql, conn)
        data_dict = {}
        for _, row in df.iterrows():
            severity_type = row["severity_type"]
            day = row["days"]
            data_dict.setdefault(severity_type, {})
            data_dict[severity_type].setdefault(day, {})
            for sev_level in [
                "extreme_districts",
                "high_districts",
                "moderate_districts",
                "low_districts",
            ]:
                val = row.get(sev_level)
                if pd.notna(val) and val.strip():
                    district_list = [d.strip() for d in val.split(",")]
                else:
                    district_list = []
                data_dict[severity_type][day][sev_level] = district_list

        severity_color = fetch_severity_colors(circle)
        return data_dict, severity_color[0]
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/fetch_district_wise_KPI_values", methods=["POST"])
@cross_origin("*")
@jwt_required()
def fetch_district_wise_KPI_values_7days():
    conn =  get_db_conn()
    try:
        data = request.get_json()  
        circle = data.get("circle")
        sql = f""" 
            select * from weatherdata.district_wise_7dayfc_severity dwds where indus_circle = '{circle}';
            """
        df = pd.read_sql_query(sql, conn)
        result = []
        for district, group in df.groupby("district"):
            district_data = {"district": district}
            for _, row in group.iterrows():
                district_data[row["days"]] = {
                    "date": row["date"],
                    "temp_min": row["temp_min"],
                    "temp_max": row["temp_max"],
                    "rain_percent": row["rain_percent"],
                    "rain_precip": row["rain_precip"],
                    "wind": row["wind"],
                    "visibility": row["visibility"],
                    "humidity": row["humidity"],
                    "temp_max_severity": row["temp_max_severity"],
                    "temp_min_severity": row["temp_min_severity"],
                    "rain_severity": row["rain_severity"],
                    "wind_severity": row["wind_severity"],
                    "visibility_severity": row["visibility_severity"],
                    "humidity_severity": row["humidity_severity"],
                    "indus_circle": row["indus_circle"],
                }
            result.append(district_data)
        final_result = []
        final_result.append({"district_wise_kpi_values": result})
        return final_result
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/generate_pdf_link", methods=["POST"])
@cross_origin("*")
@jwt_required()
def generate_pdf_link():
    try:
        payload = request.get_json()
        circle = payload.get("circle")
        date = payload.get("date")
        if not circle:
            return (
                jsonify(
                    {
                        "status": "Bad Request",
                        "message": "Circle is required for generating the PDF URL.",
                    }
                ),
                400,
            )

        if date:
            current_date = date
        else:
            current_date = datetime.now().strftime("%d %b %Y")
        filename = f"{circle} circle - {current_date}.pdf"
        encoded_filename = quote(filename)
        full_url = (
            f"https://weather.mlinfomap.com/indus/reports/pdf_circle/{encoded_filename}"
        )
        return jsonify({"status": "success", "url": full_url}), 200
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

def fetch_severity_colors(circle):
    conn =  get_db_conn()
    try:
        sql = """ 
            SELECT circle, severity_extreme_color, severity_high_color, severity_moderate_color 
            FROM weatherdata.weather_kpi_controls
            WHERE indus_circle = %s;
        """
        df = pd.read_sql_query(sql, conn, params=[circle])
        return df.to_dict(orient="records")
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

def fetch_kpi_severity_control(circle):
    conn =  get_db_conn()
    try:
        sql = f"""select * from weatherdata.weather_kpi_controls wkc
                where indus_circle = '{circle}'; """

        df = pd.read_sql_query(sql, conn)
        data = df.to_dict(orient="records")
        return data[0]
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/fetch_kpi_legend_with_color", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_legend_with_color():
    try:
        payload = request.get_json()
        circle = payload.get("circle")

        if not circle:
            return (
                jsonify({"status": "Bad Request", "message": "Circle is required."}),
                400,
            )

        kpi = fetch_kpi_severity_control(circle)
        if not kpi:
            return jsonify({"status": "error", "message": "No data found."}), 404
        circle_name = kpi.get("circle")
        parameters = [
            "temperature",
            "rainfall",
            "wind",
            "humidity",
            "visibility",
            "avalanche",
            "landslide",
            "lightning",
            "snowfall",
            "cyclone",
            "flood",
            "min_temp",
            "accu_rainfall",
        ]
        grouped_data = {"circle": circle_name}
        for param in parameters:
            grouped_data[param] = {
                "extreme": kpi.get(f"extreme_{param}"),
                "high": kpi.get(f"high_{param}"),
                "moderate": kpi.get(f"moderate_{param}"),
                "low": kpi.get(f"low_{param}"),
            }
        color_info = {
            "extreme_color": kpi.get("severity_extreme_color"),
            "high_color": kpi.get("severity_high_color"),
            "moderate_color": kpi.get("severity_moderate_color"),
            "low_color": kpi.get("severity_low_color"),
            "extreme_min_color": kpi.get("extreme_min_color"),
            "high_min_color": kpi.get("high_min_color"),
            "moderate_min_color": kpi.get("moderate_min_color"),
            "low_min_color": kpi.get("low_min_color"),
        }
        grouped_data["color"] = color_info
        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Legend with grouped colors fetched successfully.",
                    "data": grouped_data,
                }
            ),
            200,
        )
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

# fetching the all users from licensed_user_auth table
@app.route("/get_report_user_list", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_report_user_list():
    conn = get_db_conn()
    try:
        with conn.cursor() as cursor:
            query = """
                        select id, userid, "name", status, mail, mobile, indus_circle, status_activation_date, status_deactivation_date 
                        from weatherdata.master_users where team = 'indus' order by userid asc;
                        """
            cursor.execute(query)
            rows = cursor.fetchall()
            colnames = [desc[0] for desc in cursor.description]
            result = [dict(zip(colnames, row)) for row in rows]
            return jsonify({"status": "success", "data": result})
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)


@app.route("/update_report_user_status", methods=["POST"])
@cross_origin("*")
@jwt_required()
def update_report_user_status():
    conn = get_db_conn()
    try:
        payload = request.json
        userid = payload.get("id")
        status = payload.get("status")

        if not userid:
            return jsonify({"status": "error", "message": "Missing id for update"}), 400

        with conn.cursor() as cursor:
            current_datetime = datetime.now().strftime("%d %b %Y %H:%M")

            if status == "active":
                query = """
                    UPDATE weatherdata.master_users
                    SET status = %s,
                        status_activation_date = %s,
                        status_deactivation_date = NULL
                    WHERE userid = %s
                """
                params = (status, current_datetime, userid)

            else:
                query = """
                    UPDATE weatherdata.master_users
                    SET status = %s,
                        status_deactivation_date = %s
                    WHERE userid = %s
                """
                params = (status, current_datetime, userid)

            cursor.execute(query, params)
            conn.commit()
            return jsonify(
                {"status": "success", "message": "User status updated successfully"}
            )

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally: 
        release_db_conn(conn)


@app.route("/update_report_user", methods=["POST"])
@cross_origin("*")
@jwt_required()
def update_report_user():
    conn = get_db_conn()
    try:
        payload = request.get_json()
        data = payload.get("data")
        userid = payload.get("id")
        if not userid:
            return jsonify({"status": "error", "message": "Missing id for update"}), 400

        set_clause = ", ".join([f"{col} = %s" for col in data.keys()])
        values = list(data.values()) + [userid]
        with conn.cursor() as cursor:
            query = f"""
                        Update weatherdata.master_users
                        set {set_clause}
                        where userid = %s;
                    """
            cur.execute(query, values)
            conn.commit()
            return jsonify({"status": "success", "message": "User updated successfully"})

    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/add_new_report_user", methods=["POST"])
@cross_origin("*")
@jwt_required()
def add_new_report_user():
    conn = get_db_conn()
    try:
        payload = request.json
        data = payload.get("data")

        if not data:
            return jsonify({"status": "error", "message": "User data is required"}), 400
        # Required fields
        required_fields = [
            "userid",
            "name",
            "username",
            "password",
            "status",
            "team",
            "mail",
            "mobile",
            "indus_circle",
            "to_cc",
        ]

        # Validation for missing required fields
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": f"Missing required fields: {', '.join(missing_fields)}",
                    }
                ),
                400,
            )

        userid = data.get("userid")
        name = data.get("name")
        username = data.get("username")
        password = data.get("password")
        status = data.get("status")
        team = data.get("team")
        mail = data.get("mail")
        mobile = data.get("mobile")
        indus_circle = data.get("indus_circle")
        to_cc = data.get("to_cc")
        
        # Duplicate check
        with conn.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(
                "SELECT * FROM weatherdata.master_users WHERE username = %s OR mail = %s",
                (username, mail),
            )
            user = cur.fetchone()

            if user:
                return (
                    jsonify(
                        {"status": "Bad Request", "message": "User already exists"}
                    ),
                    400,
                )

        # Add status activation & deactivation timestamps based on user status
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if status == "active":
            status_activation_date = current_time
            status_deactivation_date = None
        else:
            status_activation_date = None
            status_deactivation_date = None

        # Build final INSERT query
        insert_query = """
            INSERT INTO weatherdata.master_users 
            (userid, name, username, password, status, team, mail, mobile, indus_circle, to_cc,
             status_activation_date, status_deactivation_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        values = [
            userid,
            name,
            username,
            password,
            status,
            team,
            mail,
            mobile,
            indus_circle,
            to_cc,
            status_activation_date,
            status_deactivation_date,
        ]

        with conn.cursor() as cur:
            cur.execute(insert_query, values)
            conn.commit()

        return jsonify({"status": "success", "message": "User added successfully"}), 201
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/fetch_accumulated_rainfall", methods=["POST"])
@cross_origin("*")
@jwt_required()
def fetch_accumulated_rainfall():
    conn = get_db_conn()
    try:
        payload = request.get_json()
        circle = payload.get("circle")

        if not circle:
            return jsonify({"status": "error", "message": "Circle is required"}), 400

        cur = conn.cursor(cursor_factory=RealDictCursor)
        query = """
            SELECT *
            FROM weatherdata.district_wise_accum_rainfall
            WHERE indus_circle = %s
            ORDER BY district ASC
        """

        cur.execute(query, (circle,))
        result = cur.fetchall()

        return (
            jsonify({"status": "success", "data": result, "count": len(result)}),
            200,
        )

    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": "Internal Server Error", "error": str(e)}
            ),
            500,
        )

    finally:
        release_db_conn(conn)
      
@app.route("/get_indus_circle_boundary", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_indus_circle_boundary():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        indus_circle = payload.get("circle")   
        
        sql = f"""SELECT state_ut, indus_circle, indus_zone, indus_circle_name, ST_AsText(geometry) as geometry
                    FROM weatherdata.indus_circle_geomerty where ( 'All Circle' = '{indus_circle}' or indus_circle = '{indus_circle}');"""
                    
        df = pd.read_sql(sql, conn)
        df['geometry'] = df['geometry'].apply(wkt.loads)
        gdf = gpd.GeoDataFrame(df, geometry='geometry', crs="EPSG:4326")
        
        geojson_dict = json.loads(gdf.to_json())
        
        return (
            jsonify({"status": "success", "data": geojson_dict}),
            200,
        )
        
    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": "Internal Server Error", "error": str(e)}
            ),
            500,
        )

    finally:
        release_db_conn(conn)
        
@app.route("/get_district_boundary", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_district_boundary():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        indus_circle = payload.get("circle")   
                    
        sql = f"""SELECT district, state_ut, indus_circle, indus_zone, indus_circle_name, ST_AsText(geometry) as geometry
                    FROM weatherdata.district_geometry where indus_circle is not null AND ('All Circle' = '{indus_circle}' or indus_circle = '{indus_circle}');"""
                    
        df = pd.read_sql(sql, conn)
        df['geometry'] = df['geometry'].apply(wkt.loads)
        gdf = gpd.GeoDataFrame(df, geometry='geometry', crs="EPSG:4326")
        
        geojson_dict = json.loads(gdf.to_json())
        
        return (
            jsonify({"status": "success", "data": geojson_dict}),
            200,
        )
        
    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": "Internal Server Error", "error": str(e)}
            ),
            500,
        )

    finally:
        release_db_conn(conn)
        
@app.route("/get_indus_boundary", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_indus_boundary():
    conn =  get_db_conn()
    try:
        payload = request.get_json()
        indus_circle = payload.get("circle")   
                    
        sql = f"""SELECT state_ut, indus_circle, indus_zone, ST_AsText(geometry) as geometry
                    FROM weatherdata.indus_boundary_geomerty;"""
                    
        df = pd.read_sql(sql, conn)
        df['geometry'] = df['geometry'].apply(wkt.loads)
        gdf = gpd.GeoDataFrame(df, geometry='geometry', crs="EPSG:4326")
        
        geojson_dict = json.loads(gdf.to_json())
        
        return (
            jsonify({"status": "success", "data": geojson_dict}),
            200,
        )
        
    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": "Internal Server Error", "error": str(e)}
            ),
            500,
        )

    finally:
        release_db_conn(conn)
   
@app.route("/send_usage_report", methods=["POST"])
@cross_origin("*")
@jwt_required()
def send_usage_report():
    try:
        payload = request.get_json()
        emails = payload.get("emails")
        users_data = payload.get("data") 

        if not users_data or not emails:
            return (
                jsonify({"status": "error", "message": "Missing data or emails"}),
                400,
            )

        wb = Workbook()
        wb.remove(wb.active)

        for user_rows in users_data:
            if not user_rows:
                continue

            user_name = user_rows[0].get("Name", "User")
            sheet_name = user_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Original headers
            raw_headers = list(user_rows[0].keys())
            formatted_headers = []

            for h in raw_headers:
                if h == "Login_Date_Time":
                    formatted_headers.append("Login Date & Time")
                elif h == "Logout_Date_Time":
                    formatted_headers.append("Logout Date & Time")
                else:
                    formatted_headers.append(h.replace("_", " "))

            ws.append(formatted_headers)
            for col_num, header in enumerate(formatted_headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                
            for row in user_rows:
                row_values = []
                for key in raw_headers:
                    value = row.get(key, "")

                    if isinstance(value, list):
                        value = ", ".join(value)

                    row_values.append(value)

                ws.append(row_values)

            # Auto column fit
            for col_num, header in enumerate(formatted_headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(header))

                for cell in ws[column_letter]:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))

                ws.column_dimensions[column_letter].width = max_length + 3

        folder_path = os.path.join(os.getcwd(), "usage_report")
        os.makedirs(folder_path, exist_ok=True)
        temp_file = os.path.join(folder_path, "usage_report.xlsx")
        if os.path.exists(temp_file):
            os.remove(temp_file)
        wb.save(temp_file)

        # -------- SEND EMAIL WITH ATTACHMENT --------
        yag = yagmail.SMTP(user="post@mlinfomap.com", password="tmisxyakmbllotlw")
        yag.send(
            to=emails,
            subject=f"Usages Report Data",
            contents=[
                "Please find attached excel sheet for the Usages Report Data",
                yagmail.inline(temp_file),
                temp_file,
            ],
        )

        return jsonify(
            {"status": "success", "message": "Usages report sent successfully."}
        )
    except Exception as e:
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500

@app.route("/get_india_level_districts", methods=["POST"])
@cross_origin("*")
# @jwt_required()
def get_india_level_districts():
    conn = get_db_conn()
    try:
        cur = conn.cursor(cursor_factory=DictCursor)
        # Fetch rows from last 24 hours + convert geom to GeoJSON
        query = """
            SELECT 
                *,
                ST_AsGeoJSON(geom) AS geojson_geom
            FROM weatherdata.act_warning1
            WHERE insert_at >= NOW() - INTERVAL '24 HOURS'
        """
        cur.execute(query)
        rows = cur.fetchall()

        features = []
        for row in rows:
            row_dict = dict(row)

            # Convert geometry to JSON
            geom_json = json.loads(row_dict.pop("geojson_geom"))

            # Remove original geom field
            if "geom" in row_dict:
                row_dict.pop("geom")

            feature = {"type": "Feature", "geometry": geom_json, "properties": row_dict}
            features.append(feature)
        geojson = {"type": "FeatureCollection", "features": features}
        return jsonify({"status": "success", "data": geojson}), 200
    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": "Internal Server Error", "error": str(e)}
            ),
            500,
        )

    finally:
        release_db_conn(conn)

@app.route("/insert_cyclone_file", methods=["POST"])
@cross_origin("*")
@jwt_required()
def insert_cyclone_file():
    try:
        uploaded_files = request.files
        upload_time = request.form.get("upload_time")

        if not uploaded_files or not upload_time:
            return (
                jsonify({"status": "error", "message": "Files or upload_time missing"}),
                400,
            )

        conn = get_db_conn()
        with conn.cursor() as cursor:
            for file_key, file in uploaded_files.items():
                data = json.loads(file.read())

                if "features" not in data:
                    continue

                # Detect file source type
                if "point" in file_key.lower():
                    data_type = "point"
                elif "cone" in file_key.lower():
                    data_type = "cone"
                elif "buffer" in file_key.lower():
                    data_type = "buffer"
                else:
                    data_type = "unknown"

                # Insert every feature row
                for feature in data["features"]:
                    insert_query = """
                        INSERT INTO weatherdata.cyclone_data_from_uploaded_file 
                        (data_type, properties, geometry, upload_time)
                        VALUES (%s, %s, %s, %s)
                    """
                    cursor.execute(
                        insert_query,
                        (
                            data_type,
                            json.dumps(feature.get("properties")),
                            json.dumps(feature.get("geometry")),
                            upload_time,
                        ),
                    )

            conn.commit
            return (
                jsonify(
                    {
                        "status": "success",
                        "message": "Cyclone data successfully inserted into database",
                    }
                ),
                200,
            )

    except Exception as e:
        print("Error:", e)
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/get_cyclone_geojson", methods=["POST"])
@cross_origin("*")
@jwt_required()
def get_cyclone_geojson():
    try:
        payload = request.get_json()
        selected_time = payload.get("upload_time")

        conn = get_db_conn()
        with conn.cursor() as cursor:

            # Fetch all available timestamps (for dropdown)
            cursor.execute(
                """
                SELECT DISTINCT upload_time 
                FROM weatherdata.cyclone_data_from_uploaded_file
                ORDER BY upload_time DESC;
            """
            )
            all_times = [row[0] for row in cursor.fetchall()]
            
            if not selected_time:
                selected_time = all_times[0] if all_times else None

            if not selected_time:
                return (
                    jsonify({"status": "error", "message": "No cyclone data available"}),
                    404,
                )

            # Fetch GeoJSON of selected timestamp
            cursor.execute(
                """
                SELECT data_type, properties, geometry
                FROM weatherdata.cyclone_data_from_uploaded_file
                WHERE upload_time = %s;
            """,
                (selected_time,),
            )
            rows = cursor.fetchall()

            # Prepare categorized GeoJSON skeleton
            geojson = {
                "point": {"type": "FeatureCollection", "features": []},
                "cone": {"type": "FeatureCollection", "features": []},
                "buffer": {"type": "FeatureCollection", "features": []},
            }

            for row in rows:
                data_type, props, geom = row
                feature = {
                    "type": "Feature",
                    "properties": props if props else {},
                    "geometry": geom,
                }
                # Categorize into correct feature collection
                if data_type in geojson:
                    geojson[data_type]["features"].append(feature)

            return (
                jsonify(
                    {
                        "status": "success",
                        "selected_upload_time": selected_time,
                        "all_upload_times": all_times,
                        "geojson": geojson,
                    }
                ),
                200,
            )

    except Exception as e:
        print("Error:", e)
        return jsonify({"msg": f"Internal Server error: {str(e)}"}), 500
    finally:
        release_db_conn(conn)

@app.route("/send_cyclone_report", methods=["POST"])
@cross_origin("*")
@jwt_required()
def cyclone_report_send():
    try:
        img_file = request.files.get("image")

        if not img_file:
            return jsonify({"error": "Image file missing"}), 400

        # Local API save folder
        save_folder = "./cyclone_report"
        os.makedirs(save_folder, exist_ok=True)

        # Delete previous local files
        for existing_file in os.listdir(save_folder):
            existing_file_path = os.path.join(save_folder, existing_file)
            if os.path.isfile(existing_file_path):
                os.remove(existing_file_path)

        img_path_local = os.path.join(save_folder, img_file.filename)
        img_file.save(img_path_local)

        # Save both to wwwroot folder for public access
        img_path_www = save_file_wwwroot(img_path_local, "IMAGE")

        # Public download URLs
        img_url = (
            f"https://mlinfomap.org/Weather/reports/img_cyclone/{img_file.filename}"
        )

        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Cyclone report and image uploaded successfully",
                    "image_download_url": img_url,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

def save_file_wwwroot(file_path, file_type):
    base_dir = r"C:\inetpub\wwwroot\Weather\reports"

    # Separate folders for different file types
    if file_type.upper() == "IMAGE":
        report_dir = os.path.join(base_dir, "img_cyclone")

    # Create folder if missing
    os.makedirs(report_dir, exist_ok=True)

    # Destination file
    file_name = os.path.basename(file_path)
    destination_path = os.path.join(report_dir, file_name)
    # Copy file to destination
    shutil.copyfile(file_path, destination_path)
    return destination_path


if __name__ == "__main__":
    port = int(os.environ.get("APP_PORT", 6633))
    app.run(host="0.0.0.0", port=port, debug=True)
